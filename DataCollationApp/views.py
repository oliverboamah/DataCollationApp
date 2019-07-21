from django.shortcuts import render
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from openpyxl import load_workbook


# Create your views here.
def index(request):
    return render(request, 'index.html', {})


def users(request):
    return render(request, 'users.html', {})


def upload_file(request):
    return render(request, 'upload_file.html', {})


def upload(request):
    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        fs = FileSystemStorage()
        filename = fs.save(myfile.name, myfile)
        uploaded_file_url = fs.url(filename)

        work_book = load_workbook('media/fos.xlsx')

        sheet = work_book.get_sheet_by_name(work_book.sheetnames[1])
        all_data = sheet[2:sheet.max_row]
        cell = sheet['A2']
        multi_data = []
        for i in range(sheet.max_row):
            multi_data.append([sheet['A' + str(i+2)].value, sheet['B' + str(i+2)].value,
                               sheet['C' + str(i+2)].value,
                               sheet['D' + str(i+2)].value, sheet['E' + str(i+2)].value
                               ])



        return render(request, 'upload_file.html', {
            'uploaded_file_url': multi_data
        })
    return render(request, 'upload_file.html')
