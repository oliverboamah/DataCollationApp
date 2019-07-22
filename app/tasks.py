from celery import shared_task
from openpyxl import load_workbook
from . import models


# background task to process the uploaded excel file using openpyxl
@shared_task
def upload_file_task(file_name, title):
    try:
        # open uploaded excel file's first sheet
        work_book = load_workbook('media/' + file_name)
        sheet = work_book.get_sheet_by_name(work_book.sheetnames[0])

        # array to hold individual cell entries
        data = []

        # number of rows of the excel sheet
        num_rows = sheet.max_row

        # retrieve values of cells into array
        for i in range(num_rows - 1):
            data.append([sheet['A' + str(i + 2)].value, sheet['B' + str(i + 2)].value,
                         sheet['C' + str(i + 2)].value,
                         sheet['D' + str(i + 2)].value, sheet['E' + str(i + 2)].value
                         ])

        # save metadata about the upload/submission to database
        upload_metadata = save_upload_metadata(title, file_name)

        # save all individual cell entries to database
        save_all_entries(upload_metadata, data)

        # print success status and no of rows recorded to console
        return 'Excel file uploaded successfully!, {} rows were recorded'.format(num_rows - 1)

    except Exception as e:

        # print an error message to console in case of any exception
        return e.message


# save metadata about the upload/submission to database
# returns upload metadata
def save_upload_metadata(title, file_name):
    upload_metadata = models.Upload(title=title, document_url=file_name)
    upload_metadata.publish()
    return upload_metadata


# save all individual cell entries to database
def save_all_entries(upload_metadata, data):
    for i in range(len(data)):
        individual_record = models.User(
            upload_id=upload_metadata.id,
            first_name=data[i][0],
            last_name=data[i][1],
            age=data[i][2],
            gender=data[i][3],
            address=data[i][4],
        )
        individual_record.publish()
