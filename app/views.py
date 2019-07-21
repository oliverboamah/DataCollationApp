from django.shortcuts import render
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from openpyxl import load_workbook
from . import models


# Create your views here.
def index(request):
    return render(request, 'index.html', {})


def users(request):
    return render(request, 'users.html', {})


def upload_file(request):
    return render(request, 'upload_file.html', {})


def upload(request):
    if request.method == 'POST' and request.FILES['myfile']:

        # save uploaded excel file to storage
        myfile = request.FILES['myfile']
        fs = FileSystemStorage()
        filename = fs.save(myfile.name, myfile)
        uploaded_file_url = fs.url(filename)

        # open uploaded excel file's first sheet
        work_book = load_workbook('media/' + myfile.name)
        sheet = work_book.get_sheet_by_name(work_book.sheetnames[0])

        # retrieve values of cells into a 2D array
        data = []

        # number of rows of the excel sheet
        num_rows = sheet.max_row

        for i in range(num_rows-1):
            data.append([sheet['A' + str(i + 2)].value, sheet['B' + str(i + 2)].value,
                         sheet['C' + str(i + 2)].value,
                         sheet['D' + str(i + 2)].value, sheet['E' + str(i + 2)].value
                         ])

        # save metadata about the upload/submission to database
        upload_metadata = models.Upload(title='Title', document_url=filename)
        upload_metadata.publish()

        # save individual personal records to database
        for i in range(len(data)):
            individual_record = models.User(
                upload_id=upload_metadata.id,
                first_name=data[i][0],
                last_name=data[i][1],
                age=data[i][2],
                gender=data[i][3],
                address=data[i][4],
            )

            individual_record.publish()

        return render(request, 'upload_file.html', {
            'uploaded_file_url': upload_metadata.id
        })
    return render(request, 'upload_file.html')
